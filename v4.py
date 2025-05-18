import cv2
import os
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
import insightface
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
import pickle
from insightface.app import FaceAnalysis


# === Config ===
SUBJECT_NAME = input("Enter subject name: ").strip().title()
SAVE_INTERVAL = 10
COOLDOWN_PERIOD = timedelta(hours=1)


model = FaceAnalysis(name='antelopev2')
model.prepare(ctx_id=0, det_size=(1280, 1280), det_thresh=0.4)

# === Load Pretrained Embeddings ===
def load_pretrained_embeddings(embedding_path="embeddings.pkl"):
    if not os.path.exists(embedding_path):
        raise FileNotFoundError(f"Embedding file '{embedding_path}' does not exist... please run generate_embeddings.py first.")
    with open(embedding_path, "rb") as f:
        known_embed = pickle.load(f)
    print(f"[🔁] Loaded {len(known_embed)} embeddings from {embedding_path}")
    return known_embed

known_embeddings = load_pretrained_embeddings("embeddings.pkl")

# === Attendance Excel Setup ===
excel_folder = "./attendance sheets"
os.makedirs(excel_folder, exist_ok=True)
date_str = datetime.now().strftime("%Y-%m-%d")
excel_path = f"{excel_folder}/attendance_{date_str}.xlsx"

if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = SUBJECT_NAME
    ws.append(["Name", "Time"])
    wb.save(excel_path)
    wb.close()
    print(f"[📁] Created new attendance file: {excel_path}")
else:
    wb = load_workbook(excel_path)
    if SUBJECT_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SUBJECT_NAME)
        ws.append(["Name", "Time"])
        wb.save(excel_path)
        print(f"[📝] Added new sheet '{SUBJECT_NAME}' to {excel_path}")
    wb.close()

# =ML== Load Existing Attendance if any ===
attendance = {}
try:
    df_existing = pd.read_excel(excel_path, sheet_name=SUBJECT_NAME)
    for _, row in df_existing.iterrows():
        name, timestamp = row["Name"], pd.to_datetime(row["Time"])
        # attendance[name] = timestamp
except (FileNotFoundError, InvalidFileException, ValueError):
    pass

# === Load and Process Static Image ===
img_path = "./faces/darshan.jpg"  # replace with your image path
frame = cv2.imread(img_path)

if frame is None:
    raise FileNotFoundError(f"Could not read image from {img_path}")

# Resize for better small face detection
scale_factor = 2.0
resized_frame = cv2.resize(frame, None, fx=scale_factor, fy=scale_factor)

# Detect faces on the resized image
faces = model.get(resized_frame)

for face in faces:
    emb = face.normed_embedding
    name = "n.a."

    for person, emb_list in known_embeddings.items():
        similarities = [np.dot(emb, known_emb) for known_emb in emb_list]
        max_sim = max(similarities)
        if max_sim > 0.6:
            roll_no = person  # assuming folder name = roll number
            name = f"Roll No. {roll_no}"
            now = datetime.now()
            if name not in attendance or (now - attendance[name]) > COOLDOWN_PERIOD:
                attendance[name] = now
                print(f"[✓] {name} marked at {now.strftime('%H:%M:%S')}")
            break

    # Rescale bbox to original frame size
    x1, y1, x2, y2 = (face.bbox / scale_factor).astype(int)

    # Draw thick rectangle
    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), thickness=2)

    # Draw bold text with background
    font_scale = 2.0
    text_thickness = 4
    text_size, _ = cv2.getTextSize(name, cv2.FONT_HERSHEY_SIMPLEX, font_scale, text_thickness)
    text_x, text_y = x1, y1 - 20

    # Draw background rectangle for text
    cv2.rectangle(
        frame,
        (text_x - 5, text_y - text_size[1] - 5),
        (text_x + text_size[0] + 5, text_y + 5),
        (0, 255, 0),
        -1
    )

    # Draw text (black over green bg)
    cv2.putText(
        frame,
        name,
        (text_x, text_y),
        cv2.FONT_HERSHEY_SIMPLEX,
        font_scale,
        (0, 0, 0),
        text_thickness,
        cv2.LINE_AA
    )

# === Optional: Resize image proportionally for display ===
max_width = 1280
if frame.shape[1] > max_width:
    ratio = max_width / frame.shape[1]
    frame = cv2.resize(frame, None, fx=ratio, fy=ratio)

# === Display the image ===
cv2.imshow("AttendX AI", frame)
cv2.waitKey(0)
cv2.destroyAllWindows()

# === Save Final Attendance ===
df_final = pd.DataFrame(
    [(name, timestamp.strftime("%Y-%m-%d %H:%M:%S")) for name, timestamp in attendance.items()],
    columns=["Name", "Time"]
)
with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df_final.to_excel(writer, sheet_name=SUBJECT_NAME, index=False)

print(f"[✔] Final attendance saved to '{excel_path}' under sheet '{SUBJECT_NAME}'.")
