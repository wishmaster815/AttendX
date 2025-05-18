import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import cv2
import numpy as np
import pandas as pd
import insightface
from insightface.app import FaceAnalysis
import os
from datetime import datetime, timedelta
import threading
from PIL import Image, ImageTk
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException

class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Attendance System")
        self.root.geometry("1200x900")
        
        # Prompt for subject name
        self.subject_name = simpledialog.askstring("Subject", "Enter subject name:", parent=self.root)
        if not self.subject_name:
            messagebox.showerror("Error", "Subject name is required.")
            self.root.destroy()
            return
        self.subject_name = self.subject_name.strip().title()
        
        # Attendance and cooldown
        self.COOLDOWN_PERIOD = timedelta(hours=1)
        self.attendance = {}
        
        # Excel setup
        self.excel_folder = "./attendance sheets"
        os.makedirs(self.excel_folder, exist_ok=True)
        self.date_str = datetime.now().strftime("%Y-%m-%d")
        self.excel_path = f"{self.excel_folder}/attendance_{self.date_str}.xlsx"
        self._setup_excel()
        
        # Initialize face analysis with buffalo_l model
        self.app = FaceAnalysis(name='buffalo_l', root='.')
        self.app.prepare(ctx_id=0, det_size=(640, 640))
        
        # Load embeddings
        self.embeddings = pd.read_pickle('embeddings.pkl')
        
        # Create GUI elements
        self.create_widgets()
        self.annotated_img = None
        self.tk_img = None
        
    def _setup_excel(self):
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = self.subject_name
            ws.append(["Name", "Time"])
            wb.save(self.excel_path)
            wb.close()
        else:
            wb = load_workbook(self.excel_path)
            if self.subject_name not in wb.sheetnames:
                ws = wb.create_sheet(self.subject_name)
                ws.append(["Name", "Time"])
                wb.save(self.excel_path)
            wb.close()
        # Load existing attendance
        try:
            df_existing = pd.read_excel(self.excel_path, sheet_name=self.subject_name)
            for _, row in df_existing.iterrows():
                name, timestamp = row["Name"], pd.to_datetime(row["Time"])
                self.attendance[name] = timestamp
        except (FileNotFoundError, InvalidFileException, ValueError):
            pass
    
    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        title_label = ttk.Label(main_frame, text="Attendance System", font=('Helvetica', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=2, pady=10)
        self.upload_btn = ttk.Button(main_frame, text="Upload Image", command=self.upload_file)
        self.upload_btn.grid(row=1, column=0, columnspan=2, pady=10)
        self.progress = ttk.Progressbar(main_frame, length=300, mode='determinate')
        self.progress.grid(row=2, column=0, columnspan=2, pady=10)
        self.results_text = tk.Text(main_frame, height=10, width=70)
        self.results_text.grid(row=3, column=0, columnspan=2, pady=10)
        self.status_label = ttk.Label(main_frame, text="Ready")
        self.status_label.grid(row=4, column=0, columnspan=2, pady=5)
        self.image_label = ttk.Label(main_frame)
        self.image_label.grid(row=5, column=0, columnspan=2, pady=10)
        self.save_btn = ttk.Button(main_frame, text="Save Annotated Image", command=self.save_image)
        self.save_btn.grid(row=6, column=0, columnspan=2, pady=10)
    
    def upload_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Image files", "*.jpg *.jpeg *.png")]
        )
        if file_path:
            self.process_file(file_path)
    
    def process_file(self, file_path):
        def process():
            try:
                self.status_label.config(text="Processing...")
                self.progress['value'] = 0
                self.results_text.delete(1.0, tk.END)
                img = cv2.imread(file_path)
                if img is None:
                    raise Exception("Could not read image")
                faces = self.app.get(img)
                if not faces:
                    raise Exception("No faces detected in the image")
                self.progress['value'] = 30
                results = []
                now = datetime.now()
                for face in faces:
                    bbox = face.bbox.astype(int)
                    cv2.rectangle(img, (bbox[0], bbox[1]), (bbox[2], bbox[3]), (0, 255, 0), 2)
                    embedding = np.array(face.embedding)
                    matches = []
                    for name, known_embedding in self.embeddings.items():
                        similarity = np.dot(np.array(embedding).flatten(), np.array(known_embedding).flatten())
                        if similarity > 0.5:
                            matches.append((name, similarity))
                    if matches:
                        best_match = max(matches, key=lambda x: x[1])
                        person_name = best_match[0]
                        # Draw the name/roll number above the rectangle
                        font_scale = 1.2  # Bigger text
                        text_thickness = 3  # Bolder text
                        text = str(person_name)
                        text_size, _ = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, font_scale, text_thickness)
                        text_x = bbox[0]
                        text_y = bbox[1] - 15 if bbox[1] - 15 > 15 else bbox[1] + 30
                        # Draw background rectangle for text (with more padding)
                        cv2.rectangle(
                            img,
                            (text_x - 6, text_y - text_size[1] - 6),
                            (text_x + text_size[0] + 6, text_y + 6),
                            (0, 255, 0),
                            -1
                        )
                        # Draw text (black over green bg)
                        cv2.putText(
                            img,
                            text,
                            (text_x, text_y),
                            cv2.FONT_HERSHEY_SIMPLEX,
                            font_scale,
                            (0, 0, 0),
                            text_thickness,
                            cv2.LINE_AA
                        )
                        # Cooldown logic
                        if (person_name not in self.attendance) or ((now - self.attendance[person_name]) > self.COOLDOWN_PERIOD):
                            self.attendance[person_name] = now
                            results.append(f"[✓] {person_name} marked at {now.strftime('%H:%M:%S')}")
                        else:
                            results.append(f"[!] {person_name} already marked at {self.attendance[person_name].strftime('%H:%M:%S')}")
                self.progress['value'] = 90
                if results:
                    self.results_text.insert(tk.END, f"Results for {os.path.basename(file_path)}:\n\n")
                    for result in results:
                        self.results_text.insert(tk.END, result + "\n")
                else:
                    self.results_text.insert(tk.END, "No matches found in the image.\n")
                self.annotated_img = img.copy()
                img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                pil_img = Image.fromarray(img_rgb)
                max_width, max_height = 900, 700
                pil_img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                self.tk_img = ImageTk.PhotoImage(pil_img)
                self.image_label.config(image=self.tk_img)
                self.image_label.image = self.tk_img
                # Save attendance to Excel
                self._save_attendance_to_excel()
                self.progress['value'] = 100
                self.status_label.config(text="Processing complete")
            except Exception as e:
                messagebox.showerror("Error", str(e))
                self.status_label.config(text="Error occurred")
                self.progress['value'] = 0
        thread = threading.Thread(target=process)
        thread.daemon = True
        thread.start()
    def _save_attendance_to_excel(self):
        df_final = pd.DataFrame(
            [(name, timestamp.strftime("%Y-%m-%d %H:%M:%S")) for name, timestamp in self.attendance.items()],
            columns=["Name", "Time"]
        )
        with pd.ExcelWriter(self.excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_final.to_excel(writer, sheet_name=self.subject_name, index=False)
    def save_image(self):
        if self.annotated_img is not None:
            file_path = filedialog.asksaveasfilename(defaultextension=".png")
            if file_path:
                cv2.imwrite(file_path, self.annotated_img)
                messagebox.showinfo("Saved", f"Image saved to {file_path}")
        else:
            messagebox.showerror("Error", "No image to save.")

if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceApp(root)
    root.mainloop() 